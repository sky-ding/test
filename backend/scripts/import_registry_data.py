#!/usr/bin/env python3
"""
导入已有的人力/阶段/风险数据到 registry 表（一次性手动执行，不会在服务启动时自动触发）。

支持输入格式：
1) 统一 JSON：
   {
     "manpower": {...},
     "phase": {...},
     "risk": {...}
   }
2) localStorage 导出 JSON：
   {
     "PM-tool-manpower-v1": {...},
     "PM-tool-phase-v1": {...},
     "PM-tool-risk-v1": {...}
   }
3) 顶层直接是某一个模块的数据（配合 --module 使用）。

4) Excel（.xlsx）：
   - 工作表名包含“月度执行评估（X月）” -> 导入 phase
   - 工作表名包含“人力评估（X月）” -> 导入 manpower
   - 工作表名包含“风险监控” -> 导入 risk

用法（在 backend 目录下）：
  python scripts/import_registry_data.py --file path/to/data.json --dry-run
  python scripts/import_registry_data.py --file path/to/data.json --module manpower --force
  python scripts/import_registry_data.py --file path/to/data.json --force
  python scripts/import_registry_data.py --file path/to/data.xlsx --dry-run

从 import_excel_registry.py 生成的目录一次性导入（phase.json + manpower.json + risk.json）：
  python scripts/import_registry_data.py --from-registry-dir data/registry-import --force
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from collections.abc import Iterable
from pathlib import Path
from typing import Any

BACKEND_ROOT = Path(__file__).resolve().parent.parent
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

from sqlalchemy.orm import Session  # noqa: E402

from app.db import SessionLocal, init_db  # noqa: E402
from app.models import RegistryEntry  # noqa: E402
from app.registry_store import KEY_MANPOWER, KEY_PHASE, KEY_RISK, put_json  # noqa: E402
from app.schemas import ManpowerState, PhaseState, RiskState  # noqa: E402

LOCALSTORAGE_TO_KEY = {
    "PM-tool-manpower-v1": KEY_MANPOWER,
    "PM-tool-phase-v1": KEY_PHASE,
    "PM-tool-risk-v1": KEY_RISK,
}
DIRECT_TO_KEY = {
    KEY_MANPOWER: KEY_MANPOWER,
    KEY_PHASE: KEY_PHASE,
    KEY_RISK: KEY_RISK,
    "manpower": KEY_MANPOWER,
    "phase": KEY_PHASE,
    "risk": KEY_RISK,
}
MODEL_BY_KEY = {
    KEY_MANPOWER: ManpowerState,
    KEY_PHASE: PhaseState,
    KEY_RISK: RiskState,
}
MONTH_RE = re.compile(r"（\s*(\d{1,2})\s*月\s*）")
PHASE_FIELDS = ("goal", "deliver", "highlight", "weakness", "nextNote")
SPECIAL_SET_NAMES = {"业务提效", "成本优化", "数据智能重点项目"}
SPECIAL_SUBPROJECT_NORMALIZE = {
    "业务提效": [
        ("个性化提效", "个性化提效"),
        ("大模型提效", "大模型提效"),
        ("AI智能体平台", "AI智能体平台"),
    ],
    "成本优化": [
        ("GPU精细化管理", "GPU精细化管理"),
        ("大数据计存优化", "大数据计存优化"),
        ("容器化率提升", "容器化率提升"),
        ("NLP/CV", "GPU精细化管理"),
    ],
    "数据智能重点项目": [
        ("重点业务数据建设", "重点业务数据建设"),
        ("个性化研发提效", "个性化研发提效"),
        ("核心任务稳定性保障", "核心任务稳定性保障"),
        ("货品数据Agent", "货品数据Agent"),
        ("自助分析", "自助分析和ReportX增强"),
        ("报表分析", "自助分析和ReportX增强"),
        ("ReportX", "自助分析和ReportX增强"),
        ("实验科学性优化", "实验科学性优化"),
    ],
}
SPECIAL_SUBPROJECT_EXPECTED = {
    "业务提效": ["个性化提效", "大模型提效", "AI智能体平台"],
    "成本优化": ["GPU精细化管理", "大数据计存优化", "容器化率提升"],
    "数据智能重点项目": [
        "重点业务数据建设",
        "个性化研发提效",
        "核心任务稳定性保障",
        "货品数据Agent",
        "自助分析和ReportX增强",
        "实验科学性优化",
    ],
}
EXPECTED_DEPT_GROUPS = [
    {"name": "平台与架构", "depts": ["大模型", "架构", "产品", "前端开发", "云平台", "中间件", "效能工具", "智能监控"]},
    {"name": "运维中心", "depts": ["SRE", "DBA", "监控中心"]},
    {"name": "企业服务与生产运营", "depts": ["IDC", "网络工程", "企业IT"]},
    {"name": "数据智能", "depts": ["平台开发", "数据应用", "实验平台", "数据产品"]},
]


def _to_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _to_float(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s:
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def _month_key(sheet_name: str) -> str | None:
    m = MONTH_RE.search(sheet_name)
    if not m:
        return None
    month = int(m.group(1))
    if month < 1 or month > 12:
        return None
    # 当前场景按 2026 年导入
    return f"2026-{month:02d}"


def _normalize_special_subproject(project_set: str, raw_name: str) -> str:
    rules = SPECIAL_SUBPROJECT_NORMALIZE.get(project_set, [])
    text = _to_text(raw_name)
    for key, target in rules:
        if key and key in text:
            return target
    return text


def _iter_data_rows(rows: Iterable[tuple[Any, ...]]) -> Iterable[tuple[Any, ...]]:
    for row in rows:
        # 忽略整行为空
        if not any(v is not None and str(v).strip() for v in row):
            continue
        yield row


def _parse_phase_sheet(ws: Any, month: str, phase_map: dict[tuple[str, str, str], dict[str, Any]]) -> None:
    last_program = ""
    last_project_set = ""
    for row in _iter_data_rows(ws.iter_rows(min_row=3, values_only=True)):
        program = _to_text(row[0] if len(row) > 0 else None)
        project_set = _to_text(row[1] if len(row) > 1 else None)
        project = _to_text(row[2] if len(row) > 2 else None)
        if program:
            last_program = program
        else:
            program = last_program
        if project_set:
            last_project_set = project_set
        else:
            project_set = last_project_set
        if program in SPECIAL_SET_NAMES:
            # 源表这三类常出现“项目列为空”，系统中要求“项目集=子项目”同名
            project_set = program
            project = _normalize_special_subproject(program, project_set or project or _to_text(row[1] if len(row) > 1 else None))
            # 对特殊集，子项目名称来源于第2列（原“项目集”列）
            project = _normalize_special_subproject(program, _to_text(row[1] if len(row) > 1 else None) or project)
        if not project:
            continue
        if not program:
            program = "默认项目组"
        if not project_set:
            project_set = "默认项目集"

        values = [
            _to_text(row[3] if len(row) > 3 else None),
            _to_text(row[4] if len(row) > 4 else None),
            _to_text(row[5] if len(row) > 5 else None),
            _to_text(row[6] if len(row) > 6 else None),
            _to_text(row[7] if len(row) > 7 else None),
        ]
        if not any(values):
            continue

        key = (program, project_set, project)
        item = phase_map.setdefault(
            key,
            {"program": program, "project_set": project_set, "project": project, "phaseByMonth": {}},
        )
        item["phaseByMonth"][month] = dict(zip(PHASE_FIELDS, values, strict=True))


def _parse_manpower_sheet(
    ws: Any,
    month: str,
    manpower_map: dict[tuple[str, str, str], dict[str, Any]],
    dept_names: list[str],
) -> None:
    last_program = ""
    last_project_set = ""
    # 人力 sheet 第 2 行通常是角色头，从第 4 列开始到小计前
    header_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    if header_row:
        for v in header_row[3:]:
            name = _to_text(v)
            if not name:
                continue
            if "小计" in name:
                break
            if name not in dept_names:
                dept_names.append(name)

    for row in _iter_data_rows(ws.iter_rows(min_row=3, values_only=True)):
        program = _to_text(row[0] if len(row) > 0 else None)
        project_set = _to_text(row[1] if len(row) > 1 else None)
        project = _to_text(row[2] if len(row) > 2 else None)
        if program:
            last_program = program
        else:
            program = last_program
        if project_set:
            last_project_set = project_set
        else:
            project_set = last_project_set
        if program in SPECIAL_SET_NAMES:
            # 源表这三类常出现“项目列为空”，系统中要求“项目集=子项目”同名
            project_set = program
            project = _normalize_special_subproject(program, _to_text(row[1] if len(row) > 1 else None) or project)
        if not project:
            continue
        if not program:
            program = "默认项目组"
        if not project_set:
            project_set = "默认项目集"

        # 取角色列求和作为该月总人力
        nums: list[float] = []
        for idx, v in enumerate(row[3:], start=3):
            if idx - 3 < len(dept_names):
                nums.append(_to_float(v))
        total = float(sum(nums))
        if total == 0 and program not in SPECIAL_SET_NAMES:
            continue

        key = (program, project_set, project)
        item = manpower_map.setdefault(
            key,
            {"program": program, "project_set": project_set, "project": project, "manpowerByMonth": {}},
        )
        item["manpowerByMonth"][month] = [round(total, 4)]


def _parse_risk_sheet(ws: Any) -> dict[str, Any]:
    rows: list[dict[str, Any]] = []
    for row in _iter_data_rows(ws.iter_rows(min_row=2, values_only=True)):
        solved_time = _to_text(row[7] if len(row) > 7 else None)
        risk = {
            "category": _to_text(row[0] if len(row) > 0 else None),
            "source": _to_text(row[1] if len(row) > 1 else None),
            "project": _to_text(row[2] if len(row) > 2 else None),
            "issue": _to_text(row[3] if len(row) > 3 else None),
            "solution": _to_text(row[4] if len(row) > 4 else None),
            "level": _to_text(row[5] if len(row) > 5 else None),
            "owner": _to_text(row[6] if len(row) > 6 else None),
            # 源表通常只有一个“解决时间”列，这里同时映射到 regTime/closeTime，避免页面时间字段缺失
            "regTime": solved_time,
            "closeTime": solved_time,
            "status": _to_text(row[8] if len(row) > 8 else None),
        }
        if not (risk["project"] or risk["issue"]):
            continue
        rows.append(risk)
    return {"riskRows": rows, "savedAt": None}


def _nest_projects(items: list[dict[str, Any]], key_name: str) -> list[dict[str, Any]]:
    programs: dict[str, dict[str, Any]] = {}
    for item in items:
        program = item["program"]
        project_set = item["project_set"]
        project = item["project"]
        pg = programs.setdefault(program, {"name": program, "projectSets": {}})
        ps = pg["projectSets"].setdefault(project_set, {"name": project_set, "subProjects": []})
        row = {"name": project, key_name: item[key_name]}
        if key_name == "manpowerByMonth":
            row["manpower"] = None
        ps["subProjects"].append(row)
    result = []
    for pg in programs.values():
        sets = list(pg["projectSets"].values())
        result.append({"name": pg["name"], "projectSets": sets})
    return result


def _ensure_special_project_sets(groups: list[dict[str, Any]], key_name: str) -> list[dict[str, Any]]:
    by_name = {g.get("name"): g for g in groups}
    for set_name, expected_subs in SPECIAL_SUBPROJECT_EXPECTED.items():
        program = by_name.get(set_name)
        if program is None:
            program = {"name": set_name, "projectSets": []}
            groups.append(program)
            by_name[set_name] = program

        project_sets = program.setdefault("projectSets", [])
        target_ps = None
        for ps in project_sets:
            if ps.get("name") == set_name:
                target_ps = ps
                break
        if target_ps is None:
            target_ps = {"name": set_name, "subProjects": []}
            project_sets.append(target_ps)

        existing = {sp.get("name"): sp for sp in target_ps.get("subProjects", [])}
        for sub in expected_subs:
            if sub in existing:
                continue
            row = {"name": sub, key_name: {}}
            if key_name == "manpowerByMonth":
                row["manpower"] = None
            target_ps["subProjects"].append(row)
    return groups


def _load_excel(path: Path) -> dict[str, Any]:
    try:
        import openpyxl  # type: ignore[import-not-found]
    except ImportError as exc:
        raise ValueError("缺少 openpyxl，请先执行: pip install openpyxl") from exc

    wb = openpyxl.load_workbook(path, data_only=True)
    phase_map: dict[tuple[str, str, str], dict[str, Any]] = {}
    manpower_map: dict[tuple[str, str, str], dict[str, Any]] = {}
    dept_names: list[str] = []
    risk_payload: dict[str, Any] | None = None

    for ws in wb.worksheets:
        title = _to_text(ws.title)
        if "月度执行评估" in title:
            month = _month_key(title)
            if month:
                _parse_phase_sheet(ws, month, phase_map)
            continue
        if "人力评估" in title:
            month = _month_key(title)
            if month:
                _parse_manpower_sheet(ws, month, manpower_map, dept_names)
            continue
        if "风险监控" in title:
            risk_payload = _parse_risk_sheet(ws)

    out: dict[str, Any] = {}
    if manpower_map:
        # 统一使用业务确认的部门分组顺序和归属
        dept_groups = EXPECTED_DEPT_GROUPS
        manpower_data = _nest_projects(list(manpower_map.values()), "manpowerByMonth")
        manpower_data = _ensure_special_project_sets(manpower_data, "manpowerByMonth")
        out[KEY_MANPOWER] = {
            "data": manpower_data,
            "deptGroups": dept_groups,
            "savedAt": None,
        }
    if phase_map:
        phase_data = _nest_projects(list(phase_map.values()), "phaseByMonth")
        phase_data = _ensure_special_project_sets(phase_data, "phaseByMonth")
        out[KEY_PHASE] = {
            "phaseData": phase_data,
            "savedAt": None,
        }
    if risk_payload:
        out[KEY_RISK] = risk_payload
    return out


def _load_json(path: Path) -> dict[str, Any]:
    try:
        raw = path.read_text(encoding="utf-8")
    except OSError as exc:
        raise ValueError(f"读取文件失败：{path} ({exc})") from exc
    try:
        obj = json.loads(raw)
    except json.JSONDecodeError as exc:
        raise ValueError(f"JSON 解析失败：{exc}") from exc
    if not isinstance(obj, dict):
        raise ValueError("输入 JSON 顶层必须是对象")
    return obj


def _normalize_module(module: str | None) -> str | None:
    if module is None:
        return None
    m = module.strip().lower()
    if m in ("manpower", KEY_MANPOWER):
        return KEY_MANPOWER
    if m in ("phase", KEY_PHASE):
        return KEY_PHASE
    if m in ("risk", KEY_RISK):
        return KEY_RISK
    raise ValueError(f"不支持的模块：{module}")


def _load_input(path: Path) -> dict[str, Any]:
    suffix = path.suffix.lower()
    if suffix == ".json":
        return _load_json(path)
    if suffix == ".xlsx":
        return _load_excel(path)
    raise ValueError(f"不支持的文件类型：{path.suffix}（仅支持 .json / .xlsx）")


def _load_registry_bundle_dir(dir_path: Path, module: str | None) -> dict[str, Any]:
    """读取 import_excel_registry 输出目录下的三份 JSON，组装为统一顶层键。"""
    dir_path = dir_path.expanduser().resolve()
    if not dir_path.is_dir():
        raise ValueError(f"不是目录：{dir_path}")

    files_by_key: list[tuple[str, Path]] = [
        (KEY_PHASE, dir_path / "phase.json"),
        (KEY_MANPOWER, dir_path / "manpower.json"),
        (KEY_RISK, dir_path / "risk.json"),
    ]
    if module:
        files_by_key = [(k, p) for k, p in files_by_key if k == module]
        if not files_by_key:
            raise ValueError("internal: module filter empty")
    out: dict[str, Any] = {}
    for key, path in files_by_key:
        if not path.is_file():
            raise ValueError(f"目录中缺少文件：{path.name}（路径：{path}）")
        body = _load_json(path)
        if key == KEY_PHASE and "phaseData" not in body:
            raise ValueError(f"{path.name} 缺少 phaseData")
        if key == KEY_MANPOWER and ("data" not in body or "deptGroups" not in body):
            raise ValueError(f"{path.name} 缺少 data 或 deptGroups")
        if key == KEY_RISK and "riskRows" not in body:
            raise ValueError(f"{path.name} 缺少 riskRows")
        out[key] = body
    return out


def _extract_payloads(data: dict[str, Any], module: str | None) -> dict[str, dict[str, Any]]:
    payloads: dict[str, dict[str, Any]] = {}

    if module:
        selected = data
        if module in data and isinstance(data[module], dict):
            selected = data[module]
        elif module == KEY_MANPOWER and "manpower" in data and isinstance(data["manpower"], dict):
            selected = data["manpower"]
        elif module == KEY_PHASE and "phase" in data and isinstance(data["phase"], dict):
            selected = data["phase"]
        elif module == KEY_RISK and "risk" in data and isinstance(data["risk"], dict):
            selected = data["risk"]
        elif module in LOCALSTORAGE_TO_KEY.values():
            for k, mapped in LOCALSTORAGE_TO_KEY.items():
                if mapped == module and k in data and isinstance(data[k], dict):
                    selected = data[k]
                    break
        if not isinstance(selected, dict):
            raise ValueError("模块数据必须是 JSON 对象")
        payloads[module] = selected
        return payloads

    for key, value in data.items():
        mapped = DIRECT_TO_KEY.get(str(key))
        if mapped and isinstance(value, dict):
            payloads[mapped] = value
            continue
        mapped = LOCALSTORAGE_TO_KEY.get(str(key))
        if mapped and isinstance(value, dict):
            payloads[mapped] = value

    # 顶层就是某一个模块对象时，尝试按特征识别
    if not payloads:
        if "data" in data or "deptGroups" in data:
            payloads[KEY_MANPOWER] = data
        elif "phaseData" in data:
            payloads[KEY_PHASE] = data
        elif "riskRows" in data:
            payloads[KEY_RISK] = data

    return payloads


def _validate_payloads(payloads: dict[str, dict[str, Any]]) -> dict[str, dict[str, Any]]:
    cleaned: dict[str, dict[str, Any]] = {}
    for key, payload in payloads.items():
        model = MODEL_BY_KEY[key]
        validated = model.model_validate(payload)
        cleaned[key] = validated.model_dump(exclude_none=False)
    return cleaned


def _import_to_db(
    db: Session,
    payloads: dict[str, dict[str, Any]],
    force: bool,
    dry_run: bool,
) -> tuple[list[str], list[str]]:
    written: list[str] = []
    skipped: list[str] = []
    for key, payload in payloads.items():
        exists = db.get(RegistryEntry, key) is not None
        if exists and not force:
            skipped.append(key)
            continue
        if not dry_run:
            put_json(db, key, payload)
        written.append(key)
    return written, skipped


def main() -> int:
    parser = argparse.ArgumentParser(description="Import existing registry JSON into database")
    src_group = parser.add_mutually_exclusive_group(required=True)
    src_group.add_argument("--file", type=str, default="", help="Path to input file (.json or .xlsx)")
    src_group.add_argument(
        "--from-registry-dir",
        type=str,
        default="",
        metavar="DIR",
        help="含 phase.json、manpower.json、risk.json 的目录（Excel 导入脚本输出）",
    )
    parser.add_argument(
        "--module",
        choices=["manpower", "phase", "risk"],
        help="Only import one module from the input",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Overwrite existing registry row(s) for the importing key(s)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Validate and preview import result without writing to DB",
    )
    parser.add_argument(
        "--export-json",
        default="",
        help="Optional path to save transformed payload JSON before importing",
    )
    args = parser.parse_args()

    try:
        module = _normalize_module(args.module)
        if args.from_registry_dir:
            data = _load_registry_bundle_dir(Path(args.from_registry_dir), module)
            payloads = _extract_payloads(data, None)
        else:
            src = Path(args.file).expanduser().resolve()
            if not src.is_file():
                print(f"文件不存在：{src}")
                return 1
            data = _load_input(src)
            payloads = _extract_payloads(data, module)
        if not payloads:
            print("未识别到可导入的数据。支持 manpower/phase/risk 或 PM-tool-*-v1 键。")
            return 1
        validated = _validate_payloads(payloads)
        if args.export_json:
            out_path = Path(args.export_json).expanduser().resolve()
            out_path.parent.mkdir(parents=True, exist_ok=True)
            out_path.write_text(
                json.dumps(validated, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            print(f"已导出转换结果：{out_path}")
    except ValueError as exc:
        print(f"输入校验失败：{exc}")
        return 1

    init_db()
    with SessionLocal() as db:
        written, skipped = _import_to_db(
            db=db,
            payloads=validated,
            force=args.force,
            dry_run=args.dry_run,
        )

    print(f"识别模块：{', '.join(sorted(validated.keys()))}")
    if skipped:
        print(f"跳过（数据库已有且未指定 --force）：{', '.join(sorted(skipped))}")
    if args.dry_run:
        print(f"Dry-run 成功，可写入：{', '.join(sorted(written)) if written else '无'}")
    else:
        print(f"写入完成：{', '.join(sorted(written)) if written else '无'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
