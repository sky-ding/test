#!/usr/bin/env python3
"""
从《项目管理状态月度评估*.xlsx》生成与 PM 登记工具兼容的 registry JSON（人力 / 阶段 / 风险）。

表头映射（勿按列序硬编码，以中文表头为准）见 PHASE_HEADER_TO_KEY、RISK_EXCEL_TO_FIELD。

用法（在仓库根目录）:
  python scripts/import_excel_registry.py --excel "D:/path/项目管理状态月度评估2026.xlsx" --year 2026 --out-dir backend/data/registry-import

校验: 使用 backend app.schemas 中 ManpowerState / PhaseState / RiskState 做 Pydantic 校验；
      人力行与表内「小计」列数值不一致时 stderr 告警。

写入服务端 registry 表（合并/覆盖）可在 backend 目录执行:
  python scripts/import_registry_data.py --file <合并后的 json> --dry-run
  python scripts/import_registry_data.py --file <path> --force

推送到远端 HTTP API（PUT /api/v1/*）见 backend/scripts/push_registry_to_api.py。
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

# ---------------------------------------------------------------------------
# 1) 表头中文 -> 内部 key（阶段：注意 Excel 列序与 PHASE_FIELD_KEYS 顺序不同）
# ---------------------------------------------------------------------------

PHASE_HEADER_TO_KEY: dict[str, str] = {
    "项目集": "_tree_set",
    "子项目集": "_tree_pset",
    "子项目": "_tree_proj",
    "阶段交付目标": "goal",
    "是否符合计划": "planMatch",
    "实际交付评估": "deliver",
    "执行过程分析": "highlight",
    "问题分析": "weakness",
    "改进计划": "nextNote",
}

# 风险表：Excel 列名 -> riskRows 字段（与 frontend RISK_FIELD_ORDER 一致）
RISK_EXCEL_TO_FIELD: list[tuple[str, str]] = [
    ("风险类别", "category"),
    ("风险来源", "source"),
    ("项目", "project"),
    ("问题&影响 说明", "issue"),
    ("问题&影响说明", "issue"),
    ("解决方案", "solution"),
    ("级别", "level"),
    ("跟进人", "owner"),
    ("解决时间", "closeTime"),
    ("状态", "status"),
]


def ym_key(year: int, month: int) -> str:
    return f"{year}-{month:02d}"


def parse_month_from_sheet_title(title: str) -> int | None:
    m = re.search(r"[（(]\s*(\d+)\s*月\s*[)）]", title)
    return int(m.group(1)) if m else None


def cell_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def cell_date_str(v: Any) -> str:
    if v is None or v == "":
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if len(s) >= 10 and s[4] == "-" and s[7] == "-":
        return s[:10]
    return s


def norm_risk_status(raw: Any) -> str:
    s = cell_str(raw).lower()
    if s in ("close", "closed", "done", "已关闭", "关闭", "已完成"):
        return "close"
    if s in ("hold", "pending", "搁置"):
        return "hold"
    return "open"


def forward_fill_list(seq: list[Any]) -> list[str]:
    last = ""
    out: list[str] = []
    for v in seq:
        s = cell_str(v)
        if s:
            last = s
        out.append(last)
    return out


def load_workbook(path: Path):
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as e:
        raise SystemExit("需要 openpyxl：请在 backend 虚拟环境中 pip install -r requirements.txt") from e
    return load_workbook(path, data_only=True, read_only=True)


# --- Phase tree: set -> pset -> proj -> { phaseByMonth: { ym: { keys } } } ---


def phase_tree_get(
    tree: dict[str, dict[str, dict[str, Any]]], s: str, ps: str, pr: str
) -> dict[str, Any]:
    if s not in tree:
        tree[s] = {}
    if ps not in tree[s]:
        tree[s][ps] = {}
    if pr not in tree[s][ps]:
        tree[s][ps][pr] = {"name": pr, "phaseByMonth": {}}
    return tree[s][ps][pr]


def man_tree_get(tree: dict[str, dict[str, dict[str, Any]]], s: str, ps: str, pr: str) -> dict[str, Any]:
    if s not in tree:
        tree[s] = {}
    if ps not in tree[s]:
        tree[s][ps] = {}
    if pr not in tree[s][ps]:
        tree[s][ps][pr] = {"name": pr, "manpowerByMonth": {}}
    return tree[s][ps][pr]


def phase_tree_to_phase_data(tree: dict[str, dict[str, dict[str, Any]]]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for set_name in sorted(tree.keys()):
        psets: list[dict[str, Any]] = []
        for pset_name in sorted(tree[set_name].keys()):
            subs: list[dict[str, Any]] = []
            for proj_name in sorted(tree[set_name][pset_name].keys()):
                node = tree[set_name][pset_name][proj_name]
                subs.append({"name": node["name"], "phaseByMonth": node["phaseByMonth"]})
            psets.append({"name": pset_name, "subProjects": subs})
        out.append({"name": set_name, "projectSets": psets})
    return out


def empty_phase_month_row() -> dict[str, str]:
    return {k: "" for k in ("goal", "deliver", "planMatch", "highlight", "weakness", "nextNote")}


def parse_phase_sheets(wb, year: int) -> dict[str, dict[str, dict[str, Any]]]:
    plain: dict[str, dict[str, dict[str, Any]]] = {}

    for sheet_name in wb.sheetnames:
        if "执行评估" not in sheet_name and "月度执行" not in sheet_name:
            continue
        mo = parse_month_from_sheet_title(sheet_name)
        if mo is None:
            continue
        ym = ym_key(year, mo)
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        header_row = next(rows, None)
        if not header_row:
            continue
        headers = [cell_str(x) for x in header_row]
        idx: dict[str, int] = {}
        for i, h in enumerate(headers):
            if h in PHASE_HEADER_TO_KEY:
                idx[PHASE_HEADER_TO_KEY[h]] = i
        need = {"goal", "deliver", "planMatch", "highlight", "weakness", "nextNote"}
        if not need.issubset(set(idx)):
            missing = need - set(idx)
            print(f"[warn] 阶段表 {sheet_name} 缺少列映射: {missing}", file=sys.stderr)
            continue

        carry = ["", "", ""]
        for row in rows:
            if not row:
                continue
            vals = list(row)
            while len(vals) < len(headers):
                vals.append(None)
            for i in range(3):
                t = cell_str(vals[i]) if i < len(vals) else ""
                if t:
                    carry[i] = t
            s, ps, pr = carry[0], carry[1], carry[2]
            if not pr:
                continue
            node = phase_tree_get(plain, s, ps, pr)
            if ym not in node["phaseByMonth"]:
                node["phaseByMonth"][ym] = empty_phase_month_row()
            row_obj: dict[str, str] = node["phaseByMonth"][ym]
            for key in need:
                col_i = idx[key]
                v = vals[col_i] if col_i < len(vals) else None
                row_obj[key] = cell_str(v) if v is not None else ""
    return plain


# --- Manpower: dept groups from first sheet + values per month ---


def find_subtotal_col_row1_row2(row1: list[str], row2: list[str]) -> int:
    """小计可能在第 1 行（与「人力占比」并列）或第 2 行，合并格下第 2 行常为 None。"""
    n = max(len(row1), len(row2))
    for i in range(n):
        h1 = cell_str(row1[i]) if i < len(row1) else ""
        h2 = cell_str(row2[i]) if i < len(row2) else ""
        if h1 == "小计" or h2 == "小计":
            return i
    raise ValueError("人力表中未找到「小计」列（请在表头行中包含「小计」）")


def parse_dept_groups(ws) -> tuple[list[dict[str, Any]], int]:
    """从首行分组 + 次行叶子列名解析 deptGroups；返回 (deptGroups, 小计列 0-based 下标)。"""
    rows_iter = ws.iter_rows(min_row=1, max_row=2, values_only=True)
    r1 = list(next(rows_iter))
    r2 = list(next(rows_iter))
    h1 = [cell_str(x) for x in r1]
    h2 = [cell_str(x) for x in r2]
    subtotal = find_subtotal_col_row1_row2(h1, h2)
    if subtotal < 3:
        raise ValueError("小计列位置异常")
    g_row = forward_fill_list(list(r1)[3:subtotal])
    leaves = [cell_str(r2[j]) for j in range(3, subtotal)]
    dept_groups: list[dict[str, Any]] = []
    for j in range(len(leaves)):
        leaf = leaves[j]
        if not leaf:
            continue
        gname = g_row[j] if j < len(g_row) else ""
        if not gname:
            gname = "未分组"
        if not dept_groups or dept_groups[-1]["name"] != gname:
            dept_groups.append({"name": gname, "depts": [leaf]})
        else:
            dept_groups[-1]["depts"].append(leaf)
    flat_n = sum(len(g["depts"]) for g in dept_groups)
    if flat_n == 0:
        raise ValueError("未能从人力表解析出任何部门列")
    return dept_groups, subtotal


def parse_manpower_sheet(
    ws,
    year: int,
    dept_groups: list[dict[str, Any]],
    subtotal_col: int,
    man_tree: dict[str, dict[str, dict[str, Any]]],
    ym: str,
) -> None:
    flat_n = sum(len(g["depts"]) for g in dept_groups)
    carry = ["", "", ""]
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row:
            continue
        vals = list(row)
        if len(vals) <= subtotal_col:
            continue
        for i in range(3):
            t = cell_str(vals[i]) if i < len(vals) else ""
            if t:
                carry[i] = t
        s, ps, pr = carry[0], carry[1], carry[2]
        if not pr:
            continue
        nums: list[float] = []
        for j in range(3, subtotal_col):
            v = vals[j] if j < len(vals) else None
            if v is None or v == "":
                nums.append(0.0)
            else:
                try:
                    nums.append(float(v))
                except (TypeError, ValueError):
                    nums.append(0.0)
        while len(nums) < flat_n:
            nums.append(0.0)
        nums = nums[:flat_n]
        st_val = vals[subtotal_col] if subtotal_col < len(vals) else None
        if st_val is not None and st_val != "":
            try:
                st = float(st_val)
                sm = sum(nums)
                if abs(st - sm) > 0.15:
                    print(
                        f"[warn] 人力小计与列求和不一致 {ym} {s!r}/{ps!r}/{pr!r}: 小计列={st} 部门列求和={sm:.2f}",
                        file=sys.stderr,
                    )
            except (TypeError, ValueError):
                pass
        node = man_tree_get(man_tree, s, ps, pr)
        node["manpowerByMonth"][ym] = nums


def man_tree_to_data(
    man_tree: dict[str, dict[str, dict[str, Any]]], dept_groups: list[dict[str, Any]], year: int, months_seen: list[int]
) -> list[dict[str, Any]]:
    """将人力树转为 data[]（ManpowerProgramSet 形状）。"""
    flat_n = sum(len(g["depts"]) for g in dept_groups)
    default_ym = ym_key(year, min(months_seen) if months_seen else 1)
    out: list[dict[str, Any]] = []
    for set_name in sorted(man_tree.keys()):
        pss: list[dict[str, Any]] = []
        for pset_name in sorted(man_tree[set_name].keys()):
            subs: list[dict[str, Any]] = []
            for proj_name in sorted(man_tree[set_name][pset_name].keys()):
                raw = man_tree[set_name][pset_name][proj_name]
                mb = raw.get("manpowerByMonth") or {}
                # 补齐 manpower 指针月
                man = list(mb.get(default_ym) or [0.0] * flat_n)
                while len(man) < flat_n:
                    man.append(0.0)
                man = man[:flat_n]
                subs.append({"name": proj_name, "manpowerByMonth": mb, "manpower": man})
            pss.append({"name": pset_name, "subProjects": subs})
        out.append({"name": set_name, "projectSets": pss})
    return out


def merge_phase_man_trees(
    phase_tree: dict[str, dict[str, dict[str, Any]]],
    man_tree: dict[str, dict[str, dict[str, Any]]],
) -> None:
    """双向补齐：仅阶段有的 / 仅人力有的 项目。"""
    all_keys: set[tuple[str, str, str]] = set()
    for s, psd in phase_tree.items():
        for ps, prd in psd.items():
            for pr in prd:
                all_keys.add((s, ps, pr))
    for s, psd in man_tree.items():
        for ps, prd in psd.items():
            for pr in prd:
                all_keys.add((s, ps, pr))
    for s, ps, pr in all_keys:
        phase_tree_get(phase_tree, s, ps, pr)
        man_tree_get(man_tree, s, ps, pr)


def parse_manpower_all(wb, year: int) -> tuple[list[dict[str, Any]], dict[str, dict[str, dict[str, Any]]], list[int]]:
    dept_groups: list[dict[str, Any]] | None = None
    subtotal: int | None = None
    man_tree: dict[str, dict[str, dict[str, Any]]] = {}
    months_seen: list[int] = []

    for sheet_name in wb.sheetnames:
        if "人力评估" not in sheet_name and "项目人力" not in sheet_name:
            continue
        mo = parse_month_from_sheet_title(sheet_name)
        if mo is None:
            continue
        months_seen.append(mo)
        ws = wb[sheet_name]
        if dept_groups is None:
            dept_groups, subtotal = parse_dept_groups(ws)
        assert dept_groups is not None and subtotal is not None
        parse_manpower_sheet(ws, year, dept_groups, subtotal, man_tree, ym_key(year, mo))

    if not dept_groups:
        raise ValueError("未找到任何「项目人力评估」工作表")
    return dept_groups, man_tree, sorted(set(months_seen))


def parse_risk_sheet(wb) -> list[dict[str, str]]:
    name = next((n for n in wb.sheetnames if "风险" in n), None)
    if not name:
        print("[warn] 未找到风险工作表", file=sys.stderr)
        return []
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [cell_str(x) for x in rows[0]]
    col_map: dict[str, int] = {}
    for excel_name, field in RISK_EXCEL_TO_FIELD:
        for i, h in enumerate(header):
            if h.replace(" ", "") == excel_name.replace(" ", ""):
                col_map[field] = i
                break
    need_f = {"category", "source", "project", "issue", "solution", "level", "owner", "closeTime", "status"}
    if not need_f.issubset(col_map.keys()):
        missing = need_f - set(col_map.keys())
        raise ValueError(f"风险表缺少列: {missing}，当前表头: {header}")

    out_rows: list[dict[str, str]] = []
    for row in rows[1:]:
        if not row:
            continue
        vals = list(row)
        proj = cell_str(vals[col_map["project"]]) if col_map["project"] < len(vals) else ""
        if not proj and not cell_str(vals[col_map["issue"]] if col_map["issue"] < len(vals) else ""):
            continue
        r: dict[str, str] = {
            "category": cell_str(vals[col_map["category"]]) if col_map["category"] < len(vals) else "",
            "source": cell_str(vals[col_map["source"]]) if col_map["source"] < len(vals) else "",
            "project": proj,
            "issue": cell_str(vals[col_map["issue"]]) if col_map["issue"] < len(vals) else "",
            "solution": cell_str(vals[col_map["solution"]]) if col_map["solution"] < len(vals) else "",
            "level": cell_str(vals[col_map["level"]]) if col_map["level"] < len(vals) else "",
            "owner": cell_str(vals[col_map["owner"]]) if col_map["owner"] < len(vals) else "",
            "regTime": "",
            "closeTime": cell_date_str(vals[col_map["closeTime"]]) if col_map["closeTime"] < len(vals) else "",
            "status": norm_risk_status(vals[col_map["status"]] if col_map["status"] < len(vals) else ""),
        }
        out_rows.append(r)
    return out_rows


def _count_phase_leaf_projects(phase_data: list[dict[str, Any]]) -> int:
    n = 0
    for prog in phase_data:
        for pset in prog.get("projectSets") or []:
            n += len(pset.get("subProjects") or [])
    return n


def _count_manpower_leaf_projects(data: list[dict[str, Any]]) -> int:
    n = 0
    for prog in data:
        for pset in prog.get("projectSets") or []:
            n += len(pset.get("subProjects") or [])
    return n


def validate_bundle(
    phase_data: list[dict[str, Any]],
    data: list[dict[str, Any]],
    dept_groups: list[dict[str, Any]],
    risk_rows: list[dict[str, str]],
) -> None:
    root = Path(__file__).resolve().parents[1] / "backend"
    sys.path.insert(0, str(root))
    from app.schemas import ManpowerState, PhaseState, RiskState  # noqa: E402

    PhaseState(phaseData=phase_data, savedAt=None)
    ManpowerState(data=data, deptGroups=dept_groups, savedAt=None)
    RiskState(riskRows=risk_rows, savedAt=None)
    flat_depts = sum(len(g.get("depts") or []) for g in dept_groups)
    print(
        "Pydantic 校验通过: PhaseState, ManpowerState, RiskState | "
        f"阶段叶子项目={_count_phase_leaf_projects(phase_data)}, "
        f"人力叶子项目={_count_manpower_leaf_projects(data)}, "
        f"部门列数={flat_depts}, 风险行={len(risk_rows)}",
        file=sys.stderr,
    )


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--excel", type=Path, required=True, help="xlsx 路径")
    ap.add_argument("--year", type=int, default=2026, help="数据年份（写入 ymKey）")
    ap.add_argument("--out-dir", type=Path, required=True, help="输出目录（将写入 manpower.json / phase.json / risk.json）")
    args = ap.parse_args()

    if not args.excel.is_file():
        raise SystemExit(f"文件不存在: {args.excel}")

    wb = load_workbook(args.excel)
    try:
        phase_tree = parse_phase_sheets(wb, args.year)
        dept_groups, man_tree, months_seen = parse_manpower_all(wb, args.year)
        merge_phase_man_trees(phase_tree, man_tree)
        phase_data = phase_tree_to_phase_data(phase_tree)
        data = man_tree_to_data(man_tree, dept_groups, args.year, months_seen)
        risk_rows = parse_risk_sheet(wb)
    finally:
        wb.close()

    validate_bundle(phase_data, data, dept_groups, risk_rows)

    args.out_dir.mkdir(parents=True, exist_ok=True)
    iso = datetime.utcnow().replace(microsecond=0).isoformat() + "Z"
    (args.out_dir / "phase.json").write_text(
        json.dumps({"phaseData": phase_data, "savedAt": iso}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    (args.out_dir / "manpower.json").write_text(
        json.dumps({"data": data, "deptGroups": dept_groups, "savedAt": iso}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    (args.out_dir / "risk.json").write_text(
        json.dumps({"riskRows": risk_rows, "savedAt": iso}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"已写入: {args.out_dir / 'phase.json'}, manpower.json, risk.json", file=sys.stderr)


if __name__ == "__main__":
    main()
